"""Export the dataset to a formatted workbook for Power BI / Excel."""
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parents[1] / "out"
HDR = Font(bold=True, color="FFFFFF", size=10)
FILL = PatternFill("solid", fgColor="2A78D6")
NUMS = {"std_total","std_approved","std_denied","exp_total","exp_approved","exp_denied",
        "std_appeals_total","std_appeals_overturned"}
PCTS = {"std_denial_rate","exp_denial_rate","std_appeal_overturn_rate"}

def sheet(wb, title, path, widths):
    ws = wb.create_sheet(title)
    rows = list(csv.reader(path.open()))
    for r in rows:
        ws.append(r)
    for c in range(1, len(rows[0]) + 1):
        cell = ws.cell(1, c); cell.font = HDR; cell.fill = FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = widths.get(rows[0][c-1], 14)
    hdr = rows[0]
    for ri in range(2, len(rows) + 1):
        for ci, name in enumerate(hdr, start=1):
            cell = ws.cell(ri, ci)
            if cell.value in ("", None):
                continue
            if name in NUMS:
                try: cell.value = int(float(cell.value)); cell.number_format = "#,##0"
                except ValueError: pass
            elif name in PCTS:
                try: cell.value = float(cell.value) / 100; cell.number_format = "0.00%"
                except ValueError: pass
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    return ws

wb = Workbook(); wb.remove(wb.active)
sheet(wb, "Metrics", OUT / "pa_metrics_2025.csv",
      {"plan_name": 42, "parent_org": 26, "coverage_type": 22, "source_url": 60,
       "quality_flags": 40, "extraction_note": 50, "filing_id": 20})
sheet(wb, "Validation", OUT / "validation_findings.csv",
      {"detail": 92, "rule": 34, "filing_id": 20, "severity": 10})
wb.save(OUT / "pa_metrics_2025.xlsx")
print("wrote", OUT / "pa_metrics_2025.xlsx")
