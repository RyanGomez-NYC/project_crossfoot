"""
The document store: fetch a disclosure once, keep the bytes, render it readable.

Collection has two halves and this is the mechanical one. It fetches, hashes and
converts; it never produces a filing record. Something that can read -- a person,
or a model looking at the rendered text -- takes it from there and transcribes.
Keeping the two apart is what makes the numbers auditable: the SHA-256 recorded
here is what lets someone confirm the published figures came out of the document
the payer actually served, and not out of a later, quietly different one.

    python3 -m pipeline.preauth.docs <url> [<url> ...]
    python3 -m pipeline.preauth.docs --list urls.txt
    python3 -m pipeline.preauth.docs --links <url>     # just list candidate links

Each document lands in data/raw/preauth/docs/<slug>/ with:
    original.<ext>   the bytes as delivered
    SOURCE.json      url, fetched_at, content type, sha256, size, kind
    text.txt         the readable rendering, when one could be made

Needs pdftotext (brew install poppler) for PDFs and openpyxl for workbooks; both
are optional and their absence is reported, not fatal.
"""
from __future__ import annotations

import argparse
import hashlib
import html
import json
import re
import subprocess
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
DOCS = ROOT / "data" / "raw" / "preauth" / "docs"

# This crawler says what it is. The form is the long-standing convention for a
# declared robot -- the "Mozilla/5.0 (compatible; ...)" prefix, then the name and
# a URL a webmaster can look up -- the same shape Googlebot and bingbot use. It
# is not a browser string, and it is not meant to pass for one.
#
# Measured against every host this project has successfully fetched from (159 of
# them, August 2026): 152 serve this exactly as they serve a browser. Three do
# not, all Elevance -- www.anthem.com, www.anthembluecross.com and
# www.wellpoint.com -- and they do not refuse it either. They accept the
# connection and then hold it open until it times out, every time, while the
# same request with a browser string returns in under a second.
#
# That is a payer stalling a declared crawler on a disclosure the law obliges it
# to publish, and it is exactly the kind of thing this project exists to record.
# Putting a browser string back on would make the stall disappear from the data,
# which is the one outcome worth avoiding: the timeout is recorded as a failed
# fetch and published with the rest of the misses.
UA = "Mozilla/5.0 (compatible; Crossfoot/1.0; +https://ryangomez.nyc/crossfoot)"

# A PDF yielding less than this much text is a scan whatever it claims to be.
# Real filings are dense with numbers; a page of stray header text off an
# otherwise-scanned document should not count as a successful read.
TEXT_FLOOR = 400

# What a CMS-0057-F disclosure says, in the words payers actually use. Used to
# score candidate links and to tell a metrics document from a policy page.
SIGNALS = [
    r"prior authorization", r"preauthorization", r"pre-?auth",
    r"expedited", r"standard request", r"overturn", r"appeal",
    r"turnaround", r"average time", r"denial rate", r"approved",
    r"CMS-?0057", r"interoperability",
]


def slug(url: str) -> str:
    s = re.sub(r"^https?://(www\.)?", "", url)
    s = re.sub(r"[^A-Za-z0-9]+", "-", s).strip("-").lower()
    return s[:100] or "document"


def fetch(url: str, timeout: int = 90) -> tuple[bytes, str, str]:
    req = urllib.request.Request(url, headers={
        "User-Agent": UA,
        "Accept": "text/html,application/xhtml+xml,application/pdf,*/*",
        "Accept-Language": "en-US,en;q=0.9",
    })
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read(), (r.headers.get("Content-Type") or "").split(";")[0].strip(), r.geturl()


def kind_of(url: str, ctype: str, blob: bytes) -> str:
    if blob[:5] == b"%PDF-" or ctype == "application/pdf":
        return "pdf"
    if blob[:4] == b"PK\x03\x04":
        return "xlsx" if (".xlsx" in url.lower() or "sheet" in ctype) else "zip"
    return "html"


def html_to_text(blob: bytes) -> str:
    """
    Strip a page to its text, keeping table structure.

    Not a parser -- a reducer. Table cells become tab-separated so a payer's
    metrics table survives as columns a reader can line up, which is the whole
    reason to render HTML at all here.
    """
    s = blob.decode("utf-8", "replace")
    s = re.sub(r"(?is)<(script|style|noscript|svg)\b.*?</\1>", " ", s)
    s = re.sub(r"(?i)</t[dh]>", "\t", s)
    s = re.sub(r"(?i)</(tr|p|div|li|h\d|table)>", "\n", s)
    s = re.sub(r"(?i)<br\s*/?>", "\n", s)
    s = re.sub(r"(?s)<!--.*?-->", " ", s)
    s = re.sub(r"(?s)<[^>]+>", " ", s)
    s = html.unescape(s)
    s = re.sub(r"[ \x0b\f\r]+", " ", s)
    s = re.sub(r"\n\s*\n\s*\n+", "\n\n", s)
    return "\n".join(line.strip() for line in s.splitlines())


def _pypdf_to_text(src: Path) -> str:
    """
    Pure-Python fallback when poppler is not on the machine.

    Worse than `pdftotext -layout` — it does not preserve column geometry, so a
    table can come out as a run of numbers — but it is the difference between a
    disclosure being read and being invisible. Without this, a missing system
    binary silently drops every PDF filing out of the dataset, which is what
    happened until 28 Aug 2026: 75 documents sat in the store with no text.txt
    and extract.stored() skipped every one of them.
    """
    try:
        from pypdf import PdfReader
    except ImportError:
        return ""
    try:
        reader = PdfReader(str(src))
        return "\n".join((page.extract_text() or "") for page in reader.pages)
    except Exception:  # noqa: BLE001 — a corrupt PDF is a note, not a crash
        return ""


def pdf_to_text(src: Path) -> tuple[str, list[str]]:
    notes = []
    try:
        out = subprocess.run(["pdftotext", "-layout", str(src), "-"],
                             capture_output=True, timeout=180)
        text = out.stdout.decode("utf-8", "replace")
    except FileNotFoundError:
        text = _pypdf_to_text(src)
        if not text.strip():
            return "", ["pdftotext not installed (brew install poppler) and the "
                        "pypdf fallback read nothing"]
        notes.append("pdftotext not installed; read with pypdf, so table columns "
                     "may be flattened")
    except subprocess.TimeoutExpired:
        return "", ["pdftotext timed out"]
    if len(text.strip()) < TEXT_FLOOR:
        notes.append(f"only {len(text.strip())} chars of text -- this is a scan; "
                     f"render pages with pdftoppm and read them")
    return text, notes


def xlsx_to_text(src: Path) -> tuple[str, list[str]]:
    try:
        import openpyxl
    except ImportError:
        return "", ["openpyxl not installed -- pip3 install openpyxl"]
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    parts = []
    for name in wb.sheetnames:
        ws = wb[name]
        parts.append(f"# sheet: {name}\n")
        for row in ws.iter_rows():
            cells = [str(c.value) for c in row if c.value not in (None, "")]
            if cells:
                parts.append("\t".join(cells))
        parts.append("")
    return "\n".join(parts), []


def score(text: str) -> int:
    """How much this reads like a CMS-0057-F metrics document."""
    low = text.lower()
    n = sum(1 for p in SIGNALS if re.search(p, low))
    # a metrics document is full of numbers; a policy page is not
    n += min(4, len(re.findall(r"\b\d{3,}\b", text)) // 25)
    return n


def links(url: str, blob: bytes) -> list[tuple[str, str]]:
    """Every link on a page, absolute, with its text. Used for discovery."""
    s = blob.decode("utf-8", "replace")
    out, seen = [], set()
    for m in re.finditer(r'(?is)<a\b[^>]*href="([^"]+)"[^>]*>(.*?)</a>', s):
        href = urllib.parse.urljoin(url, html.unescape(m.group(1).strip()))
        text = re.sub(r"\s+", " ", re.sub(r"(?s)<[^>]+>", " ", m.group(2))).strip()
        if href.startswith(("mailto:", "javascript:", "tel:")) or href in seen:
            continue
        seen.add(href)
        out.append((href, text[:120]))
    return out


def grab(url: str, force: bool = False) -> dict:
    dest = DOCS / slug(url)
    meta_path = dest / "SOURCE.json"
    if meta_path.exists() and not force:
        return json.loads(meta_path.read_text()) | {"cached": True}

    dest.mkdir(parents=True, exist_ok=True)
    try:
        blob, ctype, final = fetch(url)
    except urllib.error.HTTPError as e:
        meta = {"url": url, "error": f"HTTP {e.code}", "fetched_at": time.strftime("%Y-%m-%dT%H:%M:%S")}
        meta_path.write_text(json.dumps(meta, indent=2))
        return meta
    except Exception as e:                                   # noqa: BLE001
        meta = {"url": url, "error": f"{type(e).__name__}: {e}",
                "fetched_at": time.strftime("%Y-%m-%dT%H:%M:%S")}
        meta_path.write_text(json.dumps(meta, indent=2))
        return meta

    kind = kind_of(final, ctype, blob)
    ext = {"pdf": "pdf", "xlsx": "xlsx", "zip": "zip"}.get(kind, "html")
    src = dest / f"original.{ext}"
    src.write_bytes(blob)

    notes: list[str] = []
    if kind == "pdf":
        text, notes = pdf_to_text(src)
    elif kind == "xlsx":
        text, notes = xlsx_to_text(src)
    elif kind == "html":
        text = html_to_text(blob)
    else:
        text, notes = "", [f"no reader for {kind}"]
    if text:
        (dest / "text.txt").write_text(text)

    meta = {
        "url": url, "final_url": final, "fetched_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "content_type": ctype, "kind": kind, "bytes": len(blob),
        "sha256": hashlib.sha256(blob).hexdigest(),
        "text_chars": len(text), "signal_score": score(text) if text else 0,
        "notes": notes, "dir": str(dest.relative_to(ROOT)),
    }
    meta_path.write_text(json.dumps(meta, indent=2))
    return meta


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("urls", nargs="*")
    ap.add_argument("--list", help="file of URLs, one per line")
    ap.add_argument("--links", action="store_true", help="print the page's links instead of storing it")
    ap.add_argument("--force", action="store_true", help="re-fetch even if cached")
    ap.add_argument("--filter", default="", help="with --links, only links matching this regex")
    args = ap.parse_args()

    urls = list(args.urls)
    if args.list:
        urls += [ln.strip() for ln in Path(args.list).read_text().splitlines()
                 if ln.strip() and not ln.startswith("#")]
    if not urls:
        ap.error("give at least one URL")

    if args.links:
        pat = re.compile(args.filter, re.I) if args.filter else None
        for u in urls:
            blob, _, final = fetch(u)
            for href, text in links(final, blob):
                if pat and not (pat.search(href) or pat.search(text)):
                    continue
                print(f"{href}\t{text}")
        return 0

    for u in urls:
        m = grab(u, force=args.force)
        if m.get("error"):
            print(f"FAIL  {m['error']:<18} {u}")
        else:
            flag = "cached" if m.get("cached") else f"{m['kind']:<5}"
            print(f"ok    {flag} score={m.get('signal_score', 0):<3} "
                  f"{m.get('text_chars', 0):>7} chars  {m['dir']}")
            for n in m.get("notes") or []:
                print(f"      note: {n}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
