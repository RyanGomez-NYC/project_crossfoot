#!/usr/bin/env python3
"""
Turn a document the crawler could not read into something a reader can.

Three of the recorded coverage gaps are not really the same problem, and lumping
them together is why they all sat unsolved:

  spreadsheet   The numbers are there, in a .xlsx, fully machine-readable. The
                crawler simply does not open binary workbooks. MetroPlus's seven
                New York filings are the whole of this category's value.
  image         The PDF is a scan. No text layer, so nothing to parse — but the
                numbers are legible to a human, or to a model that can see.
  client        The page assembles in a browser. The document exists; a static
                fetch never sees it.

This tool does the fetching and the mechanical conversion, and stops there. It
never produces a filing record. It writes artifacts — a text grid, a page image,
a rendered DOM — and something that can read them takes it from there. Keeping
the extraction step separate is the same rule the rest of the project follows:
the model reads and transcribes, code does the arithmetic.

    python3 tools/harvest.py <url> [<url> ...]
    python3 tools/harvest.py --list urls.txt

Each document lands in harvest/<slug>/ with a SOURCE.txt recording the URL, the
fetch time, the content type, and the SHA-256 of the bytes as delivered. That
hash is the provenance anchor: it is what lets someone else confirm that the
numbers eventually published came out of the document the payer actually served,
and not out of a later, quietly different one.

Runs on a Mac, not on the web host. Needs:
    pip3 install openpyxl          # only for .xlsx
    brew install poppler           # pdftotext + pdftoppm, only for PDFs
Chrome is used for --client and is found automatically if installed.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import subprocess
import time
import urllib.error
import urllib.request
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "harvest"

UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/126.0 Safari/537.36")

# A PDF that yields less than this much text is a scan, whatever it claims.
# Real CMS filings are dense with numbers; a page or two of stray header text
# from an otherwise-scanned document should not count as a successful read.
TEXT_FLOOR = 400

CHROME_CANDIDATES = [
    "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
    "/Applications/Chromium.app/Contents/MacOS/Chromium",
    "/Applications/Microsoft Edge.app/Contents/MacOS/Microsoft Edge",
    "google-chrome", "chromium", "chromium-browser",
]


def slug(url: str) -> str:
    s = re.sub(r"^https?://(www\.)?", "", url)
    s = re.sub(r"[^A-Za-z0-9]+", "-", s).strip("-").lower()
    return s[:90] or "document"


def fetch(url: str) -> tuple[bytes, str]:
    req = urllib.request.Request(url, headers={
        "User-Agent": UA,
        "Accept": "*/*",
    })
    with urllib.request.urlopen(req, timeout=90) as r:
        return r.read(), (r.headers.get("Content-Type") or "").split(";")[0].strip()


def kind_of(url: str, ctype: str, blob: bytes) -> str:
    """What is this actually, regardless of what the URL claims."""
    if blob[:4] == b"PK\x03\x04" and (".xlsx" in url.lower() or "sheet" in ctype):
        return "xlsx"
    if blob[:5] == b"%PDF-" or ctype == "application/pdf":
        return "pdf"
    return "html"


# --------------------------------------------------------------------------
# xlsx: dump every sheet as a plain text grid.
#
# No interpretation, no header detection, no guessing which column is which.
# Cells in reading order with their coordinates, so a reader can see the actual
# shape of the sheet — including merged headers and stray notes, which is
# exactly where a naive parser would go wrong on a hand-built payer workbook.
# --------------------------------------------------------------------------
def dump_xlsx(blob: bytes, dest: Path) -> list[str]:
    try:
        import openpyxl
    except ImportError:
        return ["openpyxl not installed — run: pip3 install openpyxl"]

    src = dest / "original.xlsx"
    src.write_bytes(blob)
    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    notes = []
    for name in wb.sheetnames:
        ws = wb[name]
        lines = [f"# sheet: {name}", ""]
        for row in ws.iter_rows():
            cells = [(c.coordinate, c.value) for c in row if c.value not in (None, "")]
            if not cells:
                continue
            lines.append(" | ".join(f"{coord}={value!r}" for coord, value in cells))
        path = dest / f"sheet-{re.sub(r'[^A-Za-z0-9]+', '-', name).strip('-').lower()}.txt"
        path.write_text("\n".join(lines) + "\n")
        notes.append(f"{path.name}  ({len(lines) - 2} non-empty rows)")
    wb.close()
    return notes


# --------------------------------------------------------------------------
# pdf: try for text first, and only fall back to images if there is none.
#
# Worth doing in that order. Several documents in this project were recorded as
# "image-only" on the strength of a fetch layer returning nothing, which is not
# the same finding as the PDF having no text layer. pdftotext settles it in a
# second, and a document that turns out to be readable saves a whole extraction
# pass.
# --------------------------------------------------------------------------
def dump_pdf(blob: bytes, dest: Path, dpi: int) -> list[str]:
    src = dest / "original.pdf"
    src.write_bytes(blob)
    notes = []

    if shutil.which("pdftotext"):
        txt = dest / "text.txt"
        subprocess.run(["pdftotext", "-layout", str(src), str(txt)],
                       check=False, capture_output=True)
        n = len(txt.read_text(errors="replace").strip()) if txt.exists() else 0
        if n >= TEXT_FLOOR:
            notes.append(f"text.txt  ({n:,} characters) — this PDF has a text layer "
                         f"after all; no images needed")
            return notes
        notes.append(f"text layer holds only {n} characters — treating as a scan")
        if txt.exists() and n == 0:
            txt.unlink()
    else:
        notes.append("pdftotext not found (brew install poppler) — skipping the text check")

    if not shutil.which("pdftoppm"):
        notes.append("pdftoppm not found (brew install poppler) — cannot render pages")
        return notes

    subprocess.run(["pdftoppm", "-png", "-r", str(dpi), str(src), str(dest / "page")],
                   check=False, capture_output=True)
    pages = sorted(dest.glob("page-*.png"))
    notes.append(f"{len(pages)} page image(s) at {dpi} dpi: "
                 f"{', '.join(p.name for p in pages[:6])}"
                 f"{' …' if len(pages) > 6 else ''}")
    return notes


# --------------------------------------------------------------------------
# client-rendered: let a real browser assemble the page, then keep what it built.
#
# --dump-dom gives the DOM after scripts have run, which is the whole point —
# the static fetch that recorded these as gaps saw only the shell. Chrome's own
# headless mode is used rather than a driver library so this needs nothing
# installed that a Mac with Chrome does not already have.
# --------------------------------------------------------------------------
def dump_client(url: str, dest: Path, wait: int, binary: str | None) -> list[str]:
    chrome = binary or next((c for c in CHROME_CANDIDATES
                             if Path(c).exists() or shutil.which(c)), None)
    if not chrome:
        return ["no Chrome/Chromium found — install Chrome, or pass --client-binary"]

    html = dest / "rendered.html"
    proc = subprocess.run(
        [chrome, "--headless=new", "--disable-gpu", "--no-sandbox",
         f"--virtual-time-budget={wait * 1000}", "--dump-dom", url],
        capture_output=True, text=True, timeout=wait + 60,
    )
    html.write_text(proc.stdout)

    # A crude tag strip is deliberate. The point is to see whether numbers
    # appeared, not to produce pretty output — and a parser that "helpfully"
    # normalised the text could hide the very whitespace that makes a table
    # readable further down the line.
    text = re.sub(r"<(script|style)\b.*?</\1>", " ", proc.stdout,
                  flags=re.S | re.I)
    text = re.sub(r"<[^>]+>", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    (dest / "rendered.txt").write_text(text)

    hrefs = sorted(set(re.findall(r'href="([^"]+\.(?:pdf|xlsx|csv))"', proc.stdout, re.I)))
    if hrefs:
        (dest / "links.txt").write_text("\n".join(hrefs) + "\n")

    digits = len(re.findall(r"\b\d[\d,]{2,}\b", text))
    return [f"rendered.html ({len(proc.stdout):,} bytes), rendered.txt "
            f"({digits} number-like tokens)",
            f"links.txt — {len(hrefs)} document link(s) the static fetch never saw"
            if hrefs else "no .pdf/.xlsx links in the rendered DOM"]


def harvest(url: str, out: Path, dpi: int, wait: int, force_client: bool,
            binary: str | None = None) -> dict:
    dest = out / slug(url)
    dest.mkdir(parents=True, exist_ok=True)
    record = {"url": url, "fetched_at": time.strftime("%Y-%m-%dT%H:%M:%S%z")}

    try:
        blob, ctype = fetch(url)
    except (urllib.error.URLError, urllib.error.HTTPError, OSError) as e:
        record["error"] = f"{type(e).__name__}: {e}"
        print(f"  ! {record['error']}")
        (dest / "SOURCE.txt").write_text(json.dumps(record, indent=1) + "\n")
        return record

    record["content_type"] = ctype
    record["bytes"] = len(blob)
    record["sha256"] = hashlib.sha256(blob).hexdigest()

    kind = "html" if force_client else kind_of(url, ctype, blob)
    record["kind"] = kind

    if kind == "xlsx":
        notes = dump_xlsx(blob, dest)
    elif kind == "pdf":
        notes = dump_pdf(blob, dest, dpi)
    else:
        (dest / "original.html").write_bytes(blob)
        notes = dump_client(url, dest, wait, binary)

    record["produced"] = notes
    (dest / "SOURCE.txt").write_text(json.dumps(record, indent=1) + "\n")
    for n in notes:
        print(f"  {n}")
    return record


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("urls", nargs="*")
    ap.add_argument("--list", help="file with one URL per line (# comments allowed)")
    ap.add_argument("--out", default=str(OUT), help="output directory")
    ap.add_argument("--dpi", type=int, default=200,
                    help="page render resolution; 200 is enough for tables, "
                         "300 for small print (default 200)")
    ap.add_argument("--wait", type=int, default=8,
                    help="seconds of virtual time to let scripts run (default 8)")
    ap.add_argument("--client", action="store_true",
                    help="force browser rendering even if the URL looks like a document")
    ap.add_argument("--client-binary",
                    help="path to Chrome, if it is somewhere unusual")
    args = ap.parse_args()

    urls = list(args.urls)
    if args.list:
        for line in Path(args.list).read_text().splitlines():
            line = line.split("#")[0].strip()
            if line:
                urls.append(line)
    if not urls:
        ap.error("give at least one URL, or --list")

    out = Path(args.out)
    out.mkdir(parents=True, exist_ok=True)
    results = []
    for url in urls:
        print(f"\n{url}")
        results.append(harvest(url, out, args.dpi, args.wait, args.client,
                               args.client_binary))

    (out / "INDEX.json").write_text(json.dumps(results, indent=1) + "\n")
    ok = sum(1 for r in results if "error" not in r)
    print(f"\n{ok} of {len(results)} fetched -> {out}/")
    print("Nothing here is a filing yet. Read the artifacts, transcribe what the")
    print("document prints, and set extraction_method on every record you write.")
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
